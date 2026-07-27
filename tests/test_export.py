"""Deliverable export: multi-sheet XLSX + CSV from gate-approved records."""

from datetime import date
from pathlib import Path

import openpyxl

from fointel.export import export_dataset
from fointel.schema import (
    AuditEntry,
    Confidence,
    FamilyOfficeRecord,
    FOType,
    Provenance,
    SourceClass,
    SourceRef,
)


def _rec(fo_id: str) -> FamilyOfficeRecord:
    return FamilyOfficeRecord(
        fo_id=fo_id, name=f"{fo_id} Family Office", fo_type=FOType.SFO, fo_type_evidence="ev",
        fo_type_confidence=Confidence.HIGH, hq_country="United States", hq_phone="+1 (555) 111-2222",
        discovery_source=SourceClass.SEC_EDGAR,
        verification_sources=[SourceRef(source_class=SourceClass.SEC_EDGAR, verifies="firm facts",
                                        accessed_at=date(2026, 7, 27))],
        data_as_of=date(2026, 7, 27),
        provenance={"name": Provenance(source_class=SourceClass.SEC_EDGAR, method="submissions",
                                       checked_at=date(2026, 7, 27), confidence=Confidence.HIGH),
                    "hq_country": Provenance(source_class=SourceClass.SEC_EDGAR, method="submissions",
                                             checked_at=date(2026, 7, 27), confidence=Confidence.HIGH),
                    "hq_phone": Provenance(source_class=SourceClass.SEC_EDGAR, method="submissions",
                                           checked_at=date(2026, 7, 27), confidence=Confidence.HIGH)})


def test_export_writes_all_sheets(tmp_path: Path):
    records = [_rec("fo_a"), _rec("fo_b")]
    audit = [AuditEntry(fo_id="fo_a", field="principal_email", rejected_value="x@y.com",
                        reason="undeliverable", source_class=SourceClass.FIRM_SITE,
                        checked_at=date(2026, 7, 27))]
    result = export_dataset(records, audit, out_dir=str(tmp_path))
    assert result["records"] == 2
    assert Path(result["xlsx"]).exists() and Path(result["csv"]).exists()
    wb = openpyxl.load_workbook(result["xlsx"])
    assert set(wb.sheetnames) == {"Dataset", "Provenance", "Sources", "Audit", "Data Dictionary"}
    assert wb["Dataset"].max_row == 3          # header + 2 records
    assert wb["Audit"].max_row == 2            # header + 1 withheld value
